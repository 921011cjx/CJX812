# @Author:橘子
# @email :2315253816@qq.com
# @Time  :2022/8/01 18:01
# @File  :run.py
import requests
import openpyxl
import time

'''一、'''
web = openpyxl.load_workbook('C:\\PycharmProjects\\JK\\test_data\\test_case_api.xlsx')
sh = web["login"]
case_list = []
max_row = sh.max_row

for i in range(2,max_row+1):
    dict1=dict(
        case_id = sh.cell(row=i,column=1).value,
        url = sh.cell(row=i,column=5).value,
        data = sh.cell(row=i,column=6).value,
        expect = sh.cell(row=i,column=7).value
    )
    case_list.append(dict1)


'''二、'''
header = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}
url='http://8.129.91.152:8766/futureloan/member/login'
data={"mobile_phone":"15512345678","pwd":"12345678"}
# 调用post方法返回值
result = requests.post(url=url, json=data, headers=header).json()
print(result)


'''三、'''
def write_result(filename, sheetname, row, column, final_result,aaa):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    sh.cell(row=row,column=column).value=final_result
    sh.cell(row=row,column=column+1).value=str(aaa)
    wb.save(filename)

'''四、'''



